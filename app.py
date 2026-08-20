"""
Avant Mail Portal
-----------------------
Internal tool: upload a contact list (CSV/XLSX), compose one email with merge
fields, attach files, and send an individual personalized email to each
contact through Microsoft 365 (Microsoft Graph API).

Each message is sent one-at-a-time from the configured mailbox, so every
recipient gets a normal personal email, replies land in the sender's inbox,
and copies appear in the sender's Sent Items.

Run:  python app.py     (see README.md for Azure / Microsoft 365 setup)
"""

import base64
import csv
import html
import io
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import threading
import time
import uuid
from datetime import datetime, date

from flask import (Flask, request, redirect, url_for, render_template,
                   flash, session, jsonify, send_file, abort)

# ---------------------------------------------------------------------------
# Configuration (environment variables; see .env.example)
# ---------------------------------------------------------------------------

def _load_dotenv(path=".env"):
    """Tiny .env loader so no extra dependency is needed."""
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                os.environ.setdefault(key.strip(), value.strip())

_load_dotenv()

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(BASE_DIR, "data")
UPLOAD_DIR    = os.path.join(DATA_DIR, "attachments")
DB_PATH       = os.path.join(DATA_DIR, "portal.db")

TENANT_ID       = os.environ.get("TENANT_ID", "")
CLIENT_ID       = os.environ.get("CLIENT_ID", "")
CLIENT_SECRET   = os.environ.get("CLIENT_SECRET", "")
SENDER_EMAIL    = os.environ.get("SENDER_EMAIL", "")
SENDER_NAME     = os.environ.get("SENDER_NAME", "")
PORTAL_PASSWORD = os.environ.get("PORTAL_PASSWORD", "changeme")
DRY_RUN         = os.environ.get("DRY_RUN", "true").lower() in ("1", "true", "yes")
# Let "Send yourself a test" deliver a real email even while DRY_RUN is on
# (campaign sends stay simulated). Needs the Graph credentials filled in.
TEST_SEND_REAL  = os.environ.get("TEST_SEND_REAL", "false").lower() in ("1", "true", "yes")

# Exchange Online allows ~30 messages/minute and 10,000 recipients/day per
# mailbox. Defaults stay safely under both.
RATE_PER_MINUTE   = int(os.environ.get("RATE_PER_MINUTE", "20"))
DAILY_SEND_CAP    = int(os.environ.get("DAILY_SEND_CAP", "8000"))
MAX_ATTACH_BYTES  = 15 * 1024 * 1024  # per-email total; big messages go via upload sessions
GRAPH_SIMPLE_MAX  = int(2.5 * 1024 * 1024)  # above this, Graph's 4 MB sendMail cap looms
GRAPH_CHUNK       = 3_276_800         # upload-session chunk: multiple of 320 KiB

# Open/click tracking only works if recipients' mail apps can reach the portal.
# Set PUBLIC_URL to the portal's internet-reachable address (e.g. through a
# Cloudflare Tunnel) to enable it. Leave blank = tracking off, sending still works.
PUBLIC_URL         = os.environ.get("PUBLIC_URL", "").rstrip("/")
TRACK_OPENS        = os.environ.get("TRACK_OPENS", "true").lower() in ("1", "true", "yes")
REPLY_SCAN_MINUTES = int(os.environ.get("REPLY_SCAN_MINUTES", "15"))

def tracking_active():
    return bool(PUBLIC_URL) and TRACK_OPENS

DEFAULT_FOOTER = os.environ.get(
    "DEFAULT_FOOTER",
    "Avant Real Estate | 2875 Pomona Blvd, Pomona, CA 91768\n"
    "You are receiving this because you have worked with or inquired with Avant Real Estate. "
    "If you would prefer not to receive these emails, simply reply with \"Unsubscribe\".")

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB upload ceiling
# Pasted images ride inside the "body" form field as base64 data: URIs, and
# Werkzeug caps non-file multipart fields at 500 KB by default — raise it to
# match the request ceiling or one pasted photo 413s the whole compose form.
app.config["MAX_FORM_MEMORY_SIZE"] = 32 * 1024 * 1024


# Session cookie hardening. Secure only on Render — local dev is plain http
# and a Secure cookie there would make login silently fail.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if os.environ.get("RENDER"):
    app.config["SESSION_COOKIE_SECURE"] = True


@app.errorhandler(413)
def too_large(_e):
    flash("That upload is too large — the total request limit is 32 MB. Attachments "
          f"may total up to {MAX_ATTACH_BYTES/1024/1024:.0f} MB per email; host "
          "anything bigger online and link to it instead.", "error")
    ref = request.referrer or ""
    if ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for("dashboard"))

@app.errorhandler(404)
def not_found(_e):
    return render_template("error.html", code="404", title="Page not found",
                           message="That page doesn't exist — it may have been "
                                   "deleted, or the link is out of date."), 404

@app.errorhandler(500)
def server_error(_e):
    # The traceback has already been logged by log_exception below (or Flask).
    try:
        return render_template("error.html", code="500", title="Something went wrong",
                               message="The portal hit an unexpected error. Nothing "
                                       "was sent that shouldn't have been, and your "
                                       "saved campaigns, contacts and templates are "
                                       "safe."), 500
    except Exception:
        return "Something went wrong. Please go back and try again.", 500

@app.errorhandler(Exception)
def log_exception(e):
    """Log the full traceback (visible in Render logs), then show the friendly
    500 page instead of Werkzeug's bare 'Internal Server Error'."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e  # already handled (404/413/... keep their own handlers)
    import traceback
    print(f"[error] unhandled exception on {request.method} {request.path}:")
    traceback.print_exc()
    return server_error(e)

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

from contextlib import contextmanager

@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=15000")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def _init_wal():
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.close()

_init_wal()

def init_db():
    with db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS campaigns(
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            subject TEXT NOT NULL,
            body TEXT NOT NULL,
            is_html INTEGER DEFAULT 0,
            footer TEXT DEFAULT '',
            status TEXT DEFAULT 'draft',          -- draft/sending/paused/completed/cancelled
            created_at TEXT,
            launched_at TEXT,
            parse_stats TEXT DEFAULT '{}'
        );
        CREATE TABLE IF NOT EXISTS recipients(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id TEXT NOT NULL,
            email TEXT NOT NULL,
            name TEXT DEFAULT '',
            first_name TEXT DEFAULT '',
            company TEXT DEFAULT '',
            status TEXT DEFAULT 'pending',        -- pending/sent/failed/skipped
            error TEXT DEFAULT '',
            sent_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_recip_campaign ON recipients(campaign_id, status);
        CREATE INDEX IF NOT EXISTS idx_recip_sent ON recipients(sent_at);
        CREATE TABLE IF NOT EXISTS attachments(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            size INTEGER
        );
        CREATE TABLE IF NOT EXISTS suppression(
            email TEXT PRIMARY KEY,
            reason TEXT DEFAULT '',
            added_at TEXT
        );
        CREATE TABLE IF NOT EXISTS contact_groups(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS contacts(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            email TEXT NOT NULL,
            name TEXT DEFAULT '',
            first_name TEXT DEFAULT '',
            company TEXT DEFAULT '',
            added_at TEXT,
            UNIQUE(group_id, email)
        );
        CREATE TABLE IF NOT EXISTS templates(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            subject TEXT DEFAULT '',
            body TEXT DEFAULT '',
            is_html INTEGER DEFAULT 0,
            created_at TEXT
        );
        """)
        # Safe migrations for databases created by earlier versions.
        for table, col, decl in [
            ("recipients", "token",      "TEXT"),
            ("recipients", "opened_at",  "TEXT"),
            ("recipients", "clicked_at", "TEXT"),
            ("recipients", "replied_at", "TEXT"),
            ("recipients", "bounced",    "INTEGER DEFAULT 0"),
            ("campaigns",  "parent_id",  "TEXT"),
            ("campaigns",  "audience",   "TEXT DEFAULT ''"),
            ("campaigns",  "scheduled_at", "TEXT"),
            ("campaigns",  "track",      "INTEGER DEFAULT 1"),
            ("campaigns",  "last_scan",  "TEXT"),
        ]:
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {decl}")
            except sqlite3.OperationalError:
                pass

init_db()

# ---------------------------------------------------------------------------
# Microsoft Graph
# ---------------------------------------------------------------------------

_token_cache = {"token": None, "expires": 0}
_token_lock = threading.Lock()

def graph_configured():
    return all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, SENDER_EMAIL])

def get_graph_token():
    """Client-credentials token for Microsoft Graph, cached until near expiry."""
    import msal
    with _token_lock:
        if _token_cache["token"] and time.time() < _token_cache["expires"] - 120:
            return _token_cache["token"]
        authority = f"https://login.microsoftonline.com/{TENANT_ID}"
        client = msal.ConfidentialClientApplication(
            CLIENT_ID, authority=authority, client_credential=CLIENT_SECRET)
        result = client.acquire_token_for_client(
            scopes=["https://graph.microsoft.com/.default"])
        if "access_token" not in result:
            raise RuntimeError(result.get("error_description", "Could not get Graph token"))
        _token_cache["token"] = result["access_token"]
        _token_cache["expires"] = time.time() + int(result.get("expires_in", 3600))
        return _token_cache["token"]

def _graph_error(resp):
    """(handled, err) for a non-success Graph response; sleeps on throttle."""
    if resp.status_code == 429:  # throttled — wait and signal retry
        wait = int(resp.headers.get("Retry-After", "60"))
        time.sleep(min(wait, 300))
        return "RETRY"
    try:
        detail = resp.json().get("error", {}).get("message", resp.text[:200])
    except Exception:
        detail = resp.text[:200]
    return f"HTTP {resp.status_code}: {detail}"

def send_via_graph(to_email, to_name, subject, html_body, attachments, force_real=False):
    """Send one message as SENDER_EMAIL. Returns (ok, error_message)."""
    import requests as rq
    if DRY_RUN and not force_real:
        time.sleep(0.05)  # simulate latency so progress is visible in tests
        return True, ""
    try:
        token = get_graph_token()
    except Exception as e:
        return False, f"Auth error: {e}"

    message = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html_body},
        "toRecipients": [{"emailAddress": {"address": to_email, "name": to_name or to_email}}],
    }
    if SENDER_NAME:
        message["from"] = {"emailAddress": {"address": SENDER_EMAIL, "name": SENDER_NAME}}

    attachments = attachments or []
    approx = len(html_body) + sum(len(a.get("contentBytes", "")) for a in attachments)
    if approx > GRAPH_SIMPLE_MAX:
        return _send_via_graph_large(rq, token, message, attachments)
    if attachments:
        message["attachments"] = attachments

    url = f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail"
    try:
        resp = rq.post(url,
                       headers={"Authorization": f"Bearer {token}",
                                "Content-Type": "application/json"},
                       json={"message": message, "saveToSentItems": True},
                       timeout=60)
    except Exception as e:
        return False, f"Network error: {e}"

    if resp.status_code == 202:
        return True, ""
    return False, _graph_error(resp)

def _send_via_graph_large(rq, token, message, attachments):
    """Messages over the sendMail request cap: create a draft, add each
    attachment (chunked upload session when the file itself is big), send.
    On failure the leftover draft is deleted so it doesn't litter the mailbox."""
    base = f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    msg_id = None

    def fail(err):
        if msg_id is not None:
            try:
                rq.delete(f"{base}/messages/{msg_id}", headers=headers, timeout=30)
            except Exception:
                pass
        return False, err

    try:
        resp = rq.post(base + "/messages", headers=headers, json=message, timeout=60)
        if resp.status_code not in (200, 201):
            return False, _graph_error(resp)
        msg_id = resp.json()["id"]

        for a in attachments:
            raw = base64.b64decode(a.get("contentBytes", ""))
            if len(raw) <= GRAPH_SIMPLE_MAX:
                resp = rq.post(f"{base}/messages/{msg_id}/attachments",
                               headers=headers, json=a, timeout=120)
                if resp.status_code not in (200, 201):
                    return fail(_graph_error(resp))
                continue
            item = {"attachmentType": "file", "name": a["name"], "size": len(raw)}
            if a.get("contentId"):
                item["contentId"], item["isInline"] = a["contentId"], True
            resp = rq.post(f"{base}/messages/{msg_id}/attachments/createUploadSession",
                           headers=headers, json={"AttachmentItem": item}, timeout=60)
            if resp.status_code not in (200, 201):
                return fail(_graph_error(resp))
            upload_url = resp.json()["uploadUrl"]  # pre-authenticated: no auth header
            for start in range(0, len(raw), GRAPH_CHUNK):
                chunk = raw[start:start + GRAPH_CHUNK]
                resp = rq.put(upload_url, data=chunk, timeout=300, headers={
                    "Content-Length": str(len(chunk)),
                    "Content-Range": f"bytes {start}-{start + len(chunk) - 1}/{len(raw)}"})
                if resp.status_code not in (200, 201, 202):
                    return fail(_graph_error(resp))

        resp = rq.post(f"{base}/messages/{msg_id}/send", headers=headers, timeout=60)
        if resp.status_code == 202:
            return True, ""
        return fail(_graph_error(resp))
    except Exception as e:
        return fail(f"Network error: {e}")

# ---------------------------------------------------------------------------
# Personalization + body rendering
# ---------------------------------------------------------------------------

TOKEN_RE = re.compile(r"\{\{\s*([a-zA-Z _-]+?)\s*\}\}")

def personalize(text, r):
    """Replace {{FirstName}} / {{Name}} / {{Company}} / {{Email}} tokens."""
    values = {
        "firstname": r["first_name"] or r["name"] or "there",
        "first name": r["first_name"] or r["name"] or "there",
        "first_name": r["first_name"] or r["name"] or "there",
        "name": r["name"] or r["first_name"] or "there",
        "fullname": r["name"] or r["first_name"] or "there",
        "full name": r["name"] or r["first_name"] or "there",
        "company": r["company"] or "",
        "email": r["email"],
    }
    def sub(m):
        return values.get(m.group(1).strip().lower(), m.group(0))
    return TOKEN_RE.sub(sub, text or "")

URL_IN_TEXT_RE = re.compile(r"(https?://[^\s<]+)")

def _autolink(escaped_text):
    return URL_IN_TEXT_RE.sub(r'<a href="\1">\1</a>', escaped_text)

def plain_to_html(text):
    """Escape plain text; convert paragraphs, line breaks, *,- bullets, and links."""
    lines = (text or "").replace("\r\n", "\n").split("\n")
    out, in_list = [], False
    for raw in lines:
        line = raw.rstrip()
        bullet = re.match(r"^\s*[\*\-\u2022]\s+(.*)$", line)
        if bullet:
            if not in_list:
                out.append("<ul style='margin:0 0 12px 22px;padding:0;'>")
                in_list = True
            out.append(f"<li>{_autolink(html.escape(bullet.group(1)))}</li>")
            continue
        if in_list:
            out.append("</ul>")
            in_list = False
        if line == "":
            out.append("<br>")
        else:
            out.append(_autolink(html.escape(line)) + "<br>")
    if in_list:
        out.append("</ul>")
    return "\n".join(out)

DATA_URI_IMG_RE = re.compile(r'src="data:(image/[a-zA-Z+.-]+);base64,([^"]+)"')

def extract_inline_images(html_body):
    """Pull pasted images (data: URIs) out of the body and turn them into
    Graph inline attachments referenced by cid: — most mail clients (Outlook
    especially) refuse to display data: URIs, so this is how pasted pictures
    actually show up for recipients. Returns (html, inline_attachments)."""
    inline = []
    def repl(m):
        cid = f"inline{len(inline) + 1}"
        ext = mimetypes.guess_extension(m.group(1)) or ".png"
        inline.append({
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": f"image{len(inline) + 1}{ext}",
            "contentType": m.group(1),
            "contentBytes": m.group(2),
            "contentId": cid,
            "isInline": True,
        })
        return f'src="cid:{cid}"'
    return DATA_URI_IMG_RE.sub(repl, html_body), inline

HREF_RE = re.compile(r'href="(https?://[^"]+)"')

def add_tracking(html_body, token):
    """Rewrite links through the click tracker and append the open pixel."""
    if not (tracking_active() and token):
        return html_body
    def rewrite(m):
        from urllib.parse import quote
        return f'href="{PUBLIC_URL}/c/{token}?u={quote(m.group(1), safe="")}"'
    tracked = HREF_RE.sub(rewrite, html_body)
    tracked += (f'<img src="{PUBLIC_URL}/t/{token}.gif" width="1" height="1" '
                'alt="" style="display:none;">')
    return tracked

def build_email_html(campaign, recipient, with_tracking=False):
    """Final HTML for one recipient — styled like a normal Outlook message."""
    body_src = personalize(campaign["body"], recipient)
    body_html = body_src if campaign["is_html"] else plain_to_html(body_src)
    footer_html = ""
    if campaign["footer"]:
        footer_html = ("<br><div style='font-size:9pt;color:#6b6b6b;"
                       "border-top:1px solid #dddddd;padding-top:8px;'>"
                       + plain_to_html(personalize(campaign["footer"], recipient))
                       + "</div>")
    out = ("<div style=\"font-family:Calibri,Arial,sans-serif;"
           "font-size:11pt;color:#1a1a1a;line-height:1.45;\">"
           + body_html + footer_html + "</div>")
    if with_tracking and campaign["track"]:
        out = add_tracking(out, recipient["token"])
    return out

def first_name_of(full):
    full = (full or "").strip()
    if not full:
        return ""
    return full.split()[0].strip(",").title() if full else ""

# ---------------------------------------------------------------------------
# Reply / bounce / unsubscribe detection (Microsoft Graph, Mail.Read)
# ---------------------------------------------------------------------------

BOUNCE_SENDERS = ("mailer-daemon", "postmaster", "microsoftexchange")
BOUNCE_SUBJECTS = ("undeliverable", "delivery has failed", "returned mail",
                   "delivery status notification", "mail delivery failed")

def apply_inbox_messages(conn, campaign_id, messages):
    """Match inbox messages to a campaign's recipients.
    Marks replies, auto-suppresses 'unsubscribe' replies, flags bounces.
    Pure function over message dicts so it can be tested without Graph."""
    recips = {r["email"]: dict(r) for r in conn.execute(
        "SELECT * FROM recipients WHERE campaign_id=? AND status='sent'",
        (campaign_id,))}
    if not recips:
        return {"replied": 0, "unsubscribed": 0, "bounced": 0}
    now = datetime.now().isoformat(timespec="seconds")
    result = {"replied": 0, "unsubscribed": 0, "bounced": 0}
    for m in messages:
        sender = ((m.get("from") or {}).get("emailAddress") or {}).get("address", "").lower()
        subject = (m.get("subject") or "").lower()
        preview = (m.get("bodyPreview") or "").lower()
        received = (m.get("receivedDateTime") or now)[:19]

        if any(b in sender for b in BOUNCE_SENDERS) or \
           any(subject.startswith(b) or b in subject for b in BOUNCE_SUBJECTS):
            # bodyPreview is truncated (~255 chars) and NDRs often bury the
            # failed address deeper, so search the full body too when present.
            body = (((m.get("body") or {}).get("content")) or "").lower()
            blob = subject + " " + preview + " " + body
            for email, r in recips.items():
                if email in blob and not r.get("bounced"):
                    conn.execute("UPDATE recipients SET bounced=1 WHERE id=?", (r["id"],))
                    conn.execute("INSERT OR IGNORE INTO suppression(email,reason,added_at) "
                                 "VALUES(?,?,?)", (email, "bounced", now))
                    r["bounced"] = 1
                    result["bounced"] += 1
            continue

        r = recips.get(sender)
        if not r:
            continue
        if not r.get("replied_at"):
            conn.execute("UPDATE recipients SET replied_at=? WHERE id=?",
                         (received, r["id"]))
            r["replied_at"] = received
            result["replied"] += 1
        if "unsubscribe" in subject or "unsubscribe" in preview:
            cur = conn.execute("INSERT OR IGNORE INTO suppression(email,reason,added_at) "
                               "VALUES(?,?,?)", (sender, "asked by reply", now))
            if cur.rowcount:
                result["unsubscribed"] += 1
    return result

def scan_replies(campaign_id):
    """Pull recent inbox messages from Graph and apply them. Returns (ok, info)."""
    if DRY_RUN:
        return False, ("Dry run — the portal isn't connected to the mailbox, so there "
                       "are no replies to scan.")
    if not graph_configured():
        return False, "Microsoft 365 is not connected yet."
    import requests as rq
    with db() as conn:
        camp = conn.execute("SELECT * FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
    if not camp or not camp["launched_at"]:
        return False, "This campaign hasn't been launched yet."
    since = camp["launched_at"][:19]
    try:
        token = get_graph_token()
    except Exception as e:
        return False, f"Auth error: {e}"
    url = (f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/messages"
           f"?$filter=receivedDateTime ge {since}Z"
           "&$select=from,subject,bodyPreview,body,receivedDateTime"
           "&$orderby=receivedDateTime desc&$top=50")
    messages, pages = [], 0
    while url and pages < 6:
        try:
            resp = rq.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        except Exception as e:
            return False, f"Network error: {e}"
        if resp.status_code == 403:
            return False, ("The app can send mail but can't read the inbox yet. In the "
                           "Azure app registration, add the Microsoft Graph application "
                           "permission Mail.Read and grant admin consent (README step 4).")
        if resp.status_code != 200:
            return False, f"Graph error HTTP {resp.status_code}"
        data = resp.json()
        messages += data.get("value", [])
        url = data.get("@odata.nextLink")
        pages += 1
    with db() as conn:
        result = apply_inbox_messages(conn, campaign_id, messages)
        conn.execute("UPDATE campaigns SET last_scan=? WHERE id=?",
                     (datetime.now().isoformat(timespec="seconds"), campaign_id))
    return True, (f"Checked {len(messages)} inbox messages — "
                  f"{result['replied']} new repl{'y' if result['replied']==1 else 'ies'}, "
                  f"{result['unsubscribed']} unsubscribe request(s), "
                  f"{result['bounced']} bounce(s).")

# ---------------------------------------------------------------------------
# Contact list parsing (CSV or XLSX)
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")

def _rows_from_upload(file_storage):
    filename = (file_storage.filename or "").lower()
    data = file_storage.read()
    if filename.endswith((".xlsx", ".xlsm", ".xltx")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield ["" if c is None else str(c) for c in row]
    else:  # CSV / TSV / TXT
        text = data.decode("utf-8-sig", errors="replace")
        delimiter = "\t" if ("\t" in text.splitlines()[0] if text.splitlines() else False) else ","
        for row in csv.reader(io.StringIO(text), delimiter=delimiter):
            yield row

def parse_contacts(file_storage, suppressed):
    """Return (recipients, stats). Flexible header detection, dedupe, validate."""
    if (file_storage.filename or "").lower().endswith(".xls"):
        return [], {"error": "That's an old-format Excel file (.xls). Open it in Excel "
                             "and save it as .xlsx or CSV, then upload again."}
    rows = list(_rows_from_upload(file_storage))
    if not rows:
        return [], {"error": "The file is empty."}

    header = [str(h or "").strip().lower() for h in rows[0]]

    def find(*keys, exclude=()):
        for i, h in enumerate(header):
            if any(x in h for x in exclude):
                continue
            if any(k == h or k in h for k in keys):
                return i
        return None

    email_i   = find("email", "e-mail", "mail")
    name_i    = find("full name", "customer name", "contact name", "name",
                     "customer", "contact", exclude=("first", "last", "email", "mail"))
    first_i   = find("first name", "firstname", "first")
    last_i    = find("last name", "lastname", "last", "surname")
    company_i = find("company", "organization", "organisation", "business", "firm")

    data_rows = rows[1:]
    if email_i is None:
        # No header row? Try to detect an email column from the data itself.
        for i, cell in enumerate(rows[0]):
            if EMAIL_RE.match(str(cell).strip()):
                email_i, data_rows = i, rows
                name_i = 0 if i != 0 else (1 if len(rows[0]) > 1 else None)
                break
    if email_i is None:
        return [], {"error": "Could not find an Email column. "
                             "Use headers like: Name, Email (Company optional)."}

    seen, recipients = set(), []
    stats = {"total_rows": len(data_rows), "valid": 0, "invalid": 0,
             "duplicates": 0, "suppressed": 0}

    for row in data_rows:
        def cell(i):
            return str(row[i]).strip() if (i is not None and i < len(row)) else ""
        email = cell(email_i).lower()
        if not email:
            continue
        if not EMAIL_RE.match(email):
            stats["invalid"] += 1
            continue
        if email in seen:
            stats["duplicates"] += 1
            continue
        seen.add(email)
        if email in suppressed:
            stats["suppressed"] += 1
            continue
        name = cell(name_i)
        first = cell(first_i) or first_name_of(name)
        if not name and (first or cell(last_i)):
            name = (first + " " + cell(last_i)).strip()
        recipients.append({"email": email, "name": name.title() if name.isupper() or name.islower() else name,
                           "first_name": first.title() if first else "",
                           "company": cell(company_i)})
        stats["valid"] += 1
    return recipients, stats

# ---------------------------------------------------------------------------
# Send worker (background thread)
# ---------------------------------------------------------------------------

def sent_today(conn):
    today = date.today().isoformat()
    row = conn.execute("SELECT COUNT(*) c FROM recipients WHERE status='sent' "
                       "AND sent_at LIKE ?", (today + "%",)).fetchone()
    return row["c"]

def load_attachment_payloads(conn, campaign_id):
    """Returns (payloads, missing_filenames). A missing file must be treated
    as an error by callers — never send the email without its attachment."""
    payloads, missing = [], []
    for a in conn.execute("SELECT * FROM attachments WHERE campaign_id=?", (campaign_id,)):
        try:
            with open(a["stored_path"], "rb") as f:
                content = f.read()
        except OSError:
            missing.append(a["filename"])
            continue
        ctype = mimetypes.guess_type(a["filename"])[0] or "application/octet-stream"
        payloads.append({
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": a["filename"],
            "contentType": ctype,
            "contentBytes": base64.b64encode(content).decode(),
        })
    return payloads, missing

_last_auto_scan = 0.0

def _promote_scheduled(conn):
    now = datetime.now().isoformat(timespec="seconds")
    due = conn.execute("SELECT 1 FROM campaigns WHERE status='scheduled' "
                       "AND scheduled_at <= ? LIMIT 1", (now,)).fetchone()
    if due:
        conn.execute("UPDATE campaigns SET status='sending', launched_at=? "
                     "WHERE status='scheduled' AND scheduled_at <= ?", (now, now))

def _auto_scan_replies():
    global _last_auto_scan
    if DRY_RUN or not graph_configured():
        return
    if time.time() - _last_auto_scan < REPLY_SCAN_MINUTES * 60:
        return
    _last_auto_scan = time.time()
    with db() as conn:
        ids = [r["id"] for r in conn.execute(
            "SELECT id FROM campaigns WHERE launched_at IS NOT NULL "
            "AND launched_at >= date('now','-30 days') "
            "AND status IN ('sending','paused','completed')")]
    for cid in ids:
        ok, info = scan_replies(cid)
        if not ok:
            break  # e.g. Mail.Read not granted — try again next interval

def worker_loop():
    interval = max(60.0 / max(RATE_PER_MINUTE, 1), 0.2)
    attach_cache = {}
    while True:
        try:
            with db() as conn:
                _promote_scheduled(conn)
                camp = conn.execute("SELECT * FROM campaigns WHERE status='sending' "
                                    "ORDER BY launched_at LIMIT 1").fetchone()
                cap_hit = camp and sent_today(conn) >= DAILY_SEND_CAP
                recip = None
                if camp and not cap_hit:
                    recip = conn.execute("SELECT * FROM recipients WHERE campaign_id=? "
                                         "AND status='pending' LIMIT 1",
                                         (camp["id"],)).fetchone()
                    if not recip:
                        conn.execute("UPDATE campaigns SET status='completed' WHERE id=?",
                                     (camp["id"],))
                        attach_cache.pop(camp["id"], None)
                    else:
                        # Suppression may have grown since the campaign was created.
                        sup = conn.execute("SELECT 1 FROM suppression WHERE email=?",
                                           (recip["email"],)).fetchone()
                        if sup:
                            conn.execute("UPDATE recipients SET status='skipped', "
                                         "error='On unsubscribe list' WHERE id=?",
                                         (recip["id"],))
                            recip = None
                        elif camp["id"] not in attach_cache:
                            payloads, missing = load_attachment_payloads(conn, camp["id"])
                            if missing:
                                # Pause rather than quietly send without the flyer
                                # (e.g. after a restart wiped the upload folder).
                                conn.execute("UPDATE campaigns SET status='paused' "
                                             "WHERE id=?", (camp["id"],))
                                print(f"[worker] paused campaign {camp['id']}: "
                                      f"attachment file(s) missing: {missing}")
                                recip = None
                            else:
                                attach_cache[camp["id"]] = payloads
            if not camp:
                _auto_scan_replies()
                time.sleep(2)
                continue
            if cap_hit:
                time.sleep(60)   # daily cap reached — resume after midnight
                continue
            if not recip:
                continue

            subject = personalize(camp["subject"], recip)
            body = build_email_html(camp, recip, with_tracking=True)
            body, inline = extract_inline_images(body)
            ok, err = send_via_graph(recip["email"], recip["name"], subject,
                                     body, inline + attach_cache.get(camp["id"], []))
            with db() as conn:
                if ok:
                    conn.execute("UPDATE recipients SET status='sent', sent_at=? WHERE id=?",
                                 (datetime.now().isoformat(timespec="seconds"), recip["id"]))
                elif err == "RETRY":
                    pass  # stay pending; Retry-After sleep already applied
                else:
                    conn.execute("UPDATE recipients SET status='failed', error=? WHERE id=?",
                                 (err[:300], recip["id"]))
            time.sleep(interval)
        except Exception as e:
            print("[worker] error:", e)
            time.sleep(5)

# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.context_processor
def inject_globals():
    return {"g_dry_run": DRY_RUN, "g_sender": SENDER_EMAIL,
            "g_configured": graph_configured(),
            "g_test_real": TEST_SEND_REAL and graph_configured()}

@app.before_request
def require_login():
    open_paths = {"/login", "/healthz"}
    open_prefixes = ("/static/", "/t/", "/c/")
    if request.path in open_paths or any(request.path.startswith(p) for p in open_prefixes):
        return
    if not session.get("auth"):
        return redirect(url_for("login", next=request.path))

# 1x1 transparent GIF for open tracking
_PIXEL = base64.b64decode("R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7")

@app.route("/t/<token>.gif")
def track_open(token):
    with db() as conn:
        conn.execute("UPDATE recipients SET opened_at=COALESCE(opened_at, ?) "
                     "WHERE token=?", (datetime.now().isoformat(timespec="seconds"),
                                       re.sub(r"[^a-f0-9]", "", token)))
    resp = app.response_class(_PIXEL, mimetype="image/gif")
    resp.headers["Cache-Control"] = "no-store, private"
    return resp

@app.route("/c/<token>")
def track_click(token):
    target = request.args.get("u", "")
    if not target.startswith(("http://", "https://")):
        abort(400)
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.execute("UPDATE recipients SET clicked_at=COALESCE(clicked_at, ?), "
                           "opened_at=COALESCE(opened_at, ?) WHERE token=?",
                           (now, now, re.sub(r"[^a-f0-9]", "", token)))
        if cur.rowcount == 0:
            abort(404)  # unknown token — refuse to act as an open redirector
    return redirect(target)

@app.route("/healthz")
def healthz():
    # Touch the tables the first page-loads need, so a damaged database makes
    # the service report unhealthy (Render alerts/restarts) instead of "live
    # but every real page crashes" — which is what a corrupt portal.db did.
    try:
        with db() as conn:
            conn.execute("SELECT count(*) FROM campaigns").fetchone()
            conn.execute("SELECT count(*) FROM contact_groups").fetchone()
    except Exception as e:
        return f"unhealthy: {e}", 500
    return "ok"

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        supplied = request.form.get("password") or ""
        if secrets.compare_digest(supplied, PORTAL_PASSWORD):
            session["auth"] = True
            # Only redirect within the portal — never to another site.
            nxt = request.args.get("next") or ""
            if not nxt.startswith("/") or nxt.startswith("//"):
                nxt = url_for("dashboard")
            return redirect(nxt)
        time.sleep(1.5)  # slow down password guessing
        flash("That password is not correct.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

def campaign_counts(conn, campaign_id):
    counts = {"pending": 0, "sent": 0, "failed": 0, "skipped": 0, "total": 0}
    for row in conn.execute("SELECT status, COUNT(*) c FROM recipients "
                            "WHERE campaign_id=? GROUP BY status", (campaign_id,)):
        counts[row["status"]] = row["c"]
        counts["total"] += row["c"]
    eng = conn.execute(
        "SELECT SUM(opened_at IS NOT NULL) o, SUM(clicked_at IS NOT NULL) c, "
        "SUM(replied_at IS NOT NULL) r, SUM(bounced=1) b "
        "FROM recipients WHERE campaign_id=?", (campaign_id,)).fetchone()
    counts.update({"opened": eng["o"] or 0, "clicked": eng["c"] or 0,
                   "replied": eng["r"] or 0, "bounced": eng["b"] or 0})
    return counts

@app.route("/")
def dashboard():
    with db() as conn:
        campaigns = []
        for c in conn.execute("SELECT * FROM campaigns ORDER BY created_at DESC"):
            item = dict(c)
            item["counts"] = campaign_counts(conn, c["id"])
            campaigns.append(item)
        today_count = sent_today(conn)
        sup_count = conn.execute("SELECT COUNT(*) c FROM suppression").fetchone()["c"]
    return render_template("dashboard.html", campaigns=campaigns,
                           today_count=today_count, sup_count=sup_count,
                           daily_cap=DAILY_SEND_CAP, rate=RATE_PER_MINUTE,
                           configured=graph_configured(), dry_run=DRY_RUN,
                           sender=SENDER_EMAIL)

def _list_groups(conn):
    return [dict(g) for g in conn.execute(
        "SELECT g.*, (SELECT COUNT(*) FROM contacts WHERE group_id=g.id) AS n "
        "FROM contact_groups g ORDER BY g.name")]

def _list_templates(conn):
    return [dict(t) for t in conn.execute(
        "SELECT id,name,subject,body,is_html FROM templates ORDER BY name")]

@app.route("/campaigns/new", methods=["GET", "POST"])
def new_campaign():
    with db() as conn:
        groups = _list_groups(conn)
        tpls = _list_templates(conn)
    if request.method == "GET":
        return render_template("new_campaign.html", default_footer=DEFAULT_FOOTER,
                               groups=groups, templates=tpls)

    def back():
        return render_template("new_campaign.html", default_footer=DEFAULT_FOOTER,
                               groups=groups, templates=tpls)

    name    = request.form.get("name", "").strip() or "Untitled campaign"
    subject = request.form.get("subject", "").strip()
    body    = request.form.get("body", "")
    is_html = 1 if request.form.get("is_html") else 0
    footer  = request.form.get("footer", "").strip() if request.form.get("include_footer") else ""

    if not subject or not body.strip():
        flash("A subject line and an email body are both required.", "error")
        return back()

    with db() as conn:
        suppressed = {r["email"] for r in conn.execute("SELECT email FROM suppression")}

    audience = ""
    if request.form.get("audience_source") == "group":
        gid = request.form.get("group_id", "")
        with db() as conn:
            grp = conn.execute("SELECT * FROM contact_groups WHERE id=?", (gid,)).fetchone()
            rows = conn.execute("SELECT email,name,first_name,company FROM contacts "
                                "WHERE group_id=? ORDER BY id", (gid,)).fetchall() if grp else []
        if not grp:
            flash("Choose one of your saved contact groups.", "error")
            return back()
        recipients, sup = [], 0
        for r in rows:
            if r["email"] in suppressed:
                sup += 1
                continue
            recipients.append(dict(r))
        stats = {"total_rows": len(rows), "valid": len(recipients), "invalid": 0,
                 "duplicates": 0, "suppressed": sup}
        audience = f"Group: {grp['name']}"
    else:
        contacts_file = request.files.get("contacts")
        if not contacts_file or not contacts_file.filename:
            flash("Choose a contact list file (CSV or Excel), or pick a saved group.", "error")
            return back()
        recipients, stats = parse_contacts(contacts_file, suppressed)
        if "error" in stats:
            flash(stats["error"], "error")
            return back()
        audience = f"File: {contacts_file.filename}"
    if not recipients:
        flash("No valid email addresses were found.", "error")
        return back()

    campaign_id = uuid.uuid4().hex[:12]

    # Store attachments, enforcing the Graph size ceiling (check before writing
    # anything to disk so a rejected upload leaves no orphan files behind).
    files = [f for f in request.files.getlist("attachments") if f and f.filename]
    blobs = [(f, f.read()) for f in files]
    total_bytes = sum(len(b) for _, b in blobs)
    if total_bytes > MAX_ATTACH_BYTES:
        flash(f"Attachments total {total_bytes/1024/1024:.1f} MB — the limit per email is "
              f"{MAX_ATTACH_BYTES/1024/1024:.0f} MB. Host large flyers online and link to "
              "them instead (better for spam filters too).", "error")
        return back()
    attach_rows = []
    camp_dir = os.path.join(UPLOAD_DIR, campaign_id)
    os.makedirs(camp_dir, exist_ok=True)
    for f, blob in blobs:
        safe = re.sub(r"[^A-Za-z0-9. _()-]", "_", os.path.basename(f.filename))
        path = os.path.join(camp_dir, safe)
        with open(path, "wb") as out:
            out.write(blob)
        attach_rows.append((campaign_id, safe, path, len(blob)))

    with db() as conn:
        conn.execute("INSERT INTO campaigns(id,name,subject,body,is_html,footer,status,"
                     "created_at,parse_stats,audience) VALUES(?,?,?,?,?,?,?,?,?,?)",
                     (campaign_id, name, subject, body, is_html, footer, "draft",
                      datetime.now().isoformat(timespec="seconds"), json.dumps(stats),
                      audience))
        conn.executemany("INSERT INTO recipients(campaign_id,email,name,first_name,"
                         "company,token) VALUES(?,?,?,?,?,?)",
                         [(campaign_id, r["email"], r["name"], r["first_name"],
                           r["company"], uuid.uuid4().hex)
                          for r in recipients])
        conn.executemany("INSERT INTO attachments(campaign_id,filename,stored_path,size) "
                         "VALUES(?,?,?,?)", attach_rows)
        tpl_note = ""
        if request.form.get("save_template"):
            tpl_name = request.form.get("template_name", "").strip() or name
            conn.execute("INSERT INTO templates(name,subject,body,is_html,created_at) "
                         "VALUES(?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET "
                         "subject=excluded.subject, body=excluded.body, "
                         "is_html=excluded.is_html",
                         (tpl_name, subject, body, is_html,
                          datetime.now().isoformat(timespec="seconds")))
            tpl_note = f' Saved as template "{tpl_name}" for next time.'

    flash(f"Campaign created — {stats['valid']} contacts loaded. Review it, send yourself "
          "a test, then launch." + tpl_note, "ok")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>")
def campaign_detail(campaign_id):
    with db() as conn:
        camp = conn.execute("SELECT * FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
        if not camp:
            abort(404)
        counts = campaign_counts(conn, campaign_id)
        stats = json.loads(camp["parse_stats"] or "{}")
        attachments = [dict(a) for a in conn.execute(
            "SELECT * FROM attachments WHERE campaign_id=?", (campaign_id,))]
        sample = conn.execute("SELECT * FROM recipients WHERE campaign_id=? LIMIT 1",
                              (campaign_id,)).fetchone()
        recips = [dict(r) for r in conn.execute(
            "SELECT * FROM recipients WHERE campaign_id=? ORDER BY id LIMIT 200",
            (campaign_id,))]
    preview_html = build_email_html(camp, sample) if sample else ""
    preview_subject = personalize(camp["subject"], sample) if sample else camp["subject"]
    return render_template("campaign_detail.html", c=camp, counts=counts, stats=stats,
                           attachments=attachments, recipients=recips,
                           preview_html=preview_html, preview_subject=preview_subject,
                           configured=graph_configured(), dry_run=DRY_RUN,
                           sender=SENDER_EMAIL, rate=RATE_PER_MINUTE,
                           tracking=tracking_active())

def _set_status(campaign_id, from_statuses, to_status, extra_sql=""):
    with db() as conn:
        camp = conn.execute("SELECT status FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
        if not camp or camp["status"] not in from_statuses:
            return False
        conn.execute(f"UPDATE campaigns SET status=? {extra_sql} WHERE id=?",
                     (to_status, campaign_id))
    return True

@app.route("/campaigns/<campaign_id>/launch", methods=["POST"])
def launch(campaign_id):
    if not graph_configured() and not DRY_RUN:
        flash("Microsoft 365 is not connected yet — fill in TENANT_ID, CLIENT_ID, "
              "CLIENT_SECRET and SENDER_EMAIL in .env (see README), or set DRY_RUN=true "
              "to practice.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    with db() as conn:
        conn.execute("UPDATE campaigns SET status='sending', launched_at=? "
                     "WHERE id=? AND status IN ('draft','paused','scheduled')",
                     (datetime.now().isoformat(timespec="seconds"), campaign_id))
    flash("Sending started. You can leave this page — progress continues in the background.", "ok")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>/pause", methods=["POST"])
def pause(campaign_id):
    if _set_status(campaign_id, ("sending",), "paused"):
        flash("Paused. Nothing more will send until you resume.", "ok")
    else:
        flash("This campaign isn't sending, so there's nothing to pause.", "error")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>/cancel", methods=["POST"])
def cancel(campaign_id):
    if _set_status(campaign_id, ("draft", "sending", "paused", "scheduled"), "cancelled"):
        flash("Campaign cancelled.", "ok")
    else:
        flash("This campaign is already finished or cancelled.", "error")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>/test", methods=["POST"])
def send_test(campaign_id):
    test_email = (request.form.get("test_email") or "").strip()
    if not EMAIL_RE.match(test_email):
        flash("Enter a valid email address for the test.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    with db() as conn:
        camp = conn.execute("SELECT * FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
        sample = conn.execute("SELECT * FROM recipients WHERE campaign_id=? LIMIT 1",
                              (campaign_id,)).fetchone()
        payloads, missing = load_attachment_payloads(conn, campaign_id)
    if not camp or not sample:
        abort(404)
    if missing:
        flash("Attachment file(s) missing on the server: " + ", ".join(missing)
              + ". Re-create the campaign and upload them again.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    subject = "[TEST] " + personalize(camp["subject"], sample)
    body, inline = extract_inline_images(build_email_html(camp, sample))
    force_real = TEST_SEND_REAL and graph_configured()
    ok, err = send_via_graph(test_email, "Test recipient", subject,
                             body, inline + payloads, force_real=force_real)
    if ok:
        if DRY_RUN and not force_real:
            flash(f"Test sent to {test_email} (dry run — nothing actually sent).", "ok")
        else:
            flash(f"Test sent to {test_email} — check the inbox to see exactly how it "
                  "looks, including the attachments.", "ok")
    else:
        flash(f"Test failed: {err}", "error")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>/status.json")
def status_json(campaign_id):
    with db() as conn:
        camp = conn.execute("SELECT status FROM campaigns WHERE id=?",
                            (campaign_id,)).fetchone()
        if not camp:
            abort(404)
        counts = campaign_counts(conn, campaign_id)
        today = sent_today(conn)
    remaining = counts["pending"]
    eta_min = round(remaining / max(RATE_PER_MINUTE, 1)) if remaining else 0
    return jsonify({"status": camp["status"], "counts": counts,
                    "sent_today": today, "daily_cap": DAILY_SEND_CAP,
                    "cap_reached": today >= DAILY_SEND_CAP, "eta_minutes": eta_min,
                    "tracking": tracking_active()})

@app.route("/campaigns/<campaign_id>/results.csv")
def results_csv(campaign_id):
    with db() as conn:
        rows = conn.execute("SELECT * FROM recipients "
                            "WHERE campaign_id=? ORDER BY id", (campaign_id,)).fetchall()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Email", "Name", "Company", "Status", "Error", "Sent at",
                "Opened at", "Clicked at", "Replied at", "Bounced"])
    for r in rows:
        w.writerow([r["email"], r["name"], r["company"], r["status"], r["error"],
                    r["sent_at"] or "", r["opened_at"] or "", r["clicked_at"] or "",
                    r["replied_at"] or "", "yes" if r["bounced"] else ""])
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                     as_attachment=True, download_name="campaign_results.csv",
                     mimetype="text/csv")

# ------------------------------- Contacts ---------------------------------

@app.route("/contacts", methods=["GET", "POST"])
def contacts_page():
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Choose a CSV or Excel file to upload.", "error")
            return redirect(url_for("contacts_page"))
        with db() as conn:
            rows, stats = parse_contacts(f, suppressed=set())  # stored; campaigns filter later
            if "error" in stats:
                flash(stats["error"], "error")
                return redirect(url_for("contacts_page"))
            gid = request.form.get("group_id", "")
            new_name = request.form.get("new_group", "").strip()
            if new_name:
                existing = conn.execute("SELECT id FROM contact_groups WHERE name=? "
                                        "COLLATE NOCASE", (new_name,)).fetchone()
                if existing:  # same name = same group; merge rather than duplicate
                    gid = existing["id"]
                else:
                    cur = conn.execute("INSERT INTO contact_groups(name,created_at) VALUES(?,?)",
                                       (new_name, datetime.now().isoformat(timespec="seconds")))
                    gid = cur.lastrowid
            elif not conn.execute("SELECT 1 FROM contact_groups WHERE id=?", (gid,)).fetchone():
                flash("Pick an existing group or give the new one a name.", "error")
                return redirect(url_for("contacts_page"))
            added = 0
            now = datetime.now().isoformat(timespec="seconds")
            for r in rows:
                cur = conn.execute(
                    "INSERT OR IGNORE INTO contacts(group_id,email,name,first_name,"
                    "company,added_at) VALUES(?,?,?,?,?,?)",
                    (gid, r["email"], r["name"], r["first_name"], r["company"], now))
                added += cur.rowcount
            merged = stats["valid"] - added
        flash(f"{added} new contact(s) added"
              + (f", {merged} already in the group (merged)" if merged else "")
              + (f", {stats['duplicates']} duplicate row(s) in the file" if stats['duplicates'] else "")
              + (f", {stats['invalid']} invalid address(es) skipped" if stats['invalid'] else "")
              + ".", "ok")
        return redirect(url_for("contacts_page"))
    with db() as conn:
        groups = _list_groups(conn)
    return render_template("contacts.html", groups=groups)

@app.route("/contacts/<int:gid>/delete", methods=["POST"])
def contacts_delete(gid):
    with db() as conn:
        conn.execute("DELETE FROM contacts WHERE group_id=?", (gid,))
        conn.execute("DELETE FROM contact_groups WHERE id=?", (gid,))
    flash("Group deleted. Past campaigns are unaffected.", "ok")
    return redirect(url_for("contacts_page"))

@app.route("/contacts/<int:gid>.csv")
def contacts_csv(gid):
    with db() as conn:
        rows = conn.execute("SELECT name,email,company FROM contacts WHERE group_id=? "
                            "ORDER BY name", (gid,)).fetchall()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Email", "Company"])
    for r in rows:
        w.writerow([r["name"], r["email"], r["company"]])
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                     as_attachment=True, download_name="contacts.csv", mimetype="text/csv")

# ------------------------------- Templates ---------------------------------

@app.route("/templates")
def templates_page():
    with db() as conn:
        tpls = [dict(t) for t in conn.execute(
            "SELECT * FROM templates ORDER BY name")]
    return render_template("templates.html", templates=tpls)

@app.route("/templates/<int:tid>/delete", methods=["POST"])
def template_delete(tid):
    with db() as conn:
        conn.execute("DELETE FROM templates WHERE id=?", (tid,))
    flash("Template deleted. Campaigns already created from it are unaffected.", "ok")
    return redirect(url_for("templates_page"))

# ------------------------- Follow-ups / schedule / scan --------------------

@app.route("/campaigns/<campaign_id>/followup", methods=["POST"])
def create_followup(campaign_id):
    with db() as conn:
        parent = conn.execute("SELECT * FROM campaigns WHERE id=?", (campaign_id,)).fetchone()
        if not parent:
            abort(404)
        conds, labels = ["status='sent'", "bounced=0", "replied_at IS NULL"], ["didn't reply"]
        if request.form.get("only_unopened"):
            conds.append("opened_at IS NULL")
            labels.append("didn't open")
        if request.form.get("only_unclicked"):
            conds.append("clicked_at IS NULL")
            labels.append("didn't click")
        suppressed = {r["email"] for r in conn.execute("SELECT email FROM suppression")}
        rows = [r for r in conn.execute(
            f"SELECT email,name,first_name,company FROM recipients "
            f"WHERE campaign_id=? AND {' AND '.join(conds)}", (campaign_id,))
            if r["email"] not in suppressed]
    if not rows:
        flash("No one matches that follow-up filter (everyone replied, bounced, or "
              "unsubscribed).", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))

    subject = request.form.get("subject", "").strip() or ("Following up: " + parent["subject"])
    body = request.form.get("body", "")
    if not body.strip():
        flash("Write the follow-up email body first.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    footer = parent["footer"]
    new_id = uuid.uuid4().hex[:12]
    stats = {"total_rows": len(rows), "valid": len(rows), "invalid": 0,
             "duplicates": 0, "suppressed": 0}
    with db() as conn:
        conn.execute("INSERT INTO campaigns(id,name,subject,body,is_html,footer,status,"
                     "created_at,parse_stats,audience,parent_id) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                     (new_id, "Follow-up: " + parent["name"], subject, body, 0, footer,
                      "draft", datetime.now().isoformat(timespec="seconds"),
                      json.dumps(stats), "Follow-up (" + ", ".join(labels) + ")",
                      campaign_id))
        conn.executemany("INSERT INTO recipients(campaign_id,email,name,first_name,"
                         "company,token) VALUES(?,?,?,?,?,?)",
                         [(new_id, r["email"], r["name"], r["first_name"], r["company"],
                           uuid.uuid4().hex) for r in rows])
    flash(f"Follow-up drafted for {len(rows)} contact(s) who {' and '.join(labels)}. "
          "Review, test, then launch.", "ok")
    return redirect(url_for("campaign_detail", campaign_id=new_id))

@app.route("/campaigns/<campaign_id>/schedule", methods=["POST"])
def schedule(campaign_id):
    when = request.form.get("scheduled_at", "").strip()
    try:
        dt = datetime.fromisoformat(when)
    except ValueError:
        flash("Pick a valid date and time.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    if dt <= datetime.now():
        flash("That time is in the past — pick a future time, or just launch now.", "error")
        return redirect(url_for("campaign_detail", campaign_id=campaign_id))
    with db() as conn:
        conn.execute("UPDATE campaigns SET status='scheduled', scheduled_at=? "
                     "WHERE id=? AND status IN ('draft','paused','scheduled')",
                     (dt.isoformat(timespec="seconds"), campaign_id))
    flash(f"Scheduled for {dt.strftime('%b %d, %Y at %I:%M %p')}. The portal must be "
          "running at that time.", "ok")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

@app.route("/campaigns/<campaign_id>/scan", methods=["POST"])
def scan_now(campaign_id):
    ok, info = scan_replies(campaign_id)
    flash(info, "ok" if ok else "error")
    return redirect(url_for("campaign_detail", campaign_id=campaign_id))

# ------------------------------- Unsubscribes ------------------------------

@app.route("/unsubscribes", methods=["GET", "POST"])
def unsubscribes():
    if request.method == "POST":
        raw = request.form.get("emails", "")
        added = 0
        with db() as conn:
            for token in re.split(r"[\s,;]+", raw):
                email = token.strip().lower()
                if EMAIL_RE.match(email):
                    conn.execute("INSERT OR IGNORE INTO suppression(email,reason,added_at) "
                                 "VALUES(?,?,?)", (email, "manual",
                                 datetime.now().isoformat(timespec="seconds")))
                    added += 1
        flash(f"Added {added} address(es). They will be skipped in every future campaign.", "ok")
        return redirect(url_for("unsubscribes"))
    with db() as conn:
        entries = [dict(r) for r in conn.execute(
            "SELECT * FROM suppression ORDER BY added_at DESC")]
    return render_template("suppression.html", entries=entries)

@app.route("/unsubscribes/remove", methods=["POST"])
def unsubscribe_remove():
    with db() as conn:
        conn.execute("DELETE FROM suppression WHERE email=?",
                     ((request.form.get("email") or "").lower(),))
    return redirect(url_for("unsubscribes"))

# ---------------------------------------------------------------------------

def start_worker():
    t = threading.Thread(target=worker_loop, daemon=True)
    t.start()

start_worker()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8080"))
    print(f"\n  Avant Mail Portal  ->  http://localhost:{port}")
    print(f"  Mode: {'DRY RUN (no real email is sent)' if DRY_RUN else 'LIVE via Microsoft 365'}"
          f"   Sender: {SENDER_EMAIL or '(not set)'}\n")
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False, threaded=True)
